import json
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell import Cell
from typing import List, Any

class ExcelParser:
    def __init__(self, file_path: str) -> None:
        """
        Constructor to initialize the ExcelTableParser with a file path.

        :param file_path: Path to the Excel file.
        """
        self.file_path = file_path
        self.workbook = load_workbook(file_path, data_only=True)
        self.sheet = self.workbook.active
        self.extracted_tables = []

    def cell_has_border(self, cell: Cell) -> bool:
        """
        Determine if a cell has any borders.

        :param cell: The cell to check.
        :return: True if the cell has borders, False otherwise.
        """
        return self.cell_has_row_border(cell) or self.cell_has_col_border(cell)

    def cell_has_row_border(self, cell: Cell) -> bool:
        """
        Determine if a cell has a top or bottom border.

        :param cell: The cell to check.
        :return: True if the cell has a top or bottom border, False otherwise.
        """
        return bool(cell.border.top.style or cell.border.bottom.style)

    def cell_has_col_border(self, cell: Cell) -> bool:
        """
        Determine if a cell has a left or right border.

        :param cell: The cell to check.
        :return: True if the cell has a left or right border, False otherwise.
        """
        return bool(cell.border.left.style or cell.border.right.style)

    def remove_empty_cols(self, table: List[List[Any]]) -> List[List[Any]]:
        """
        Eliminate empty columns from a table.

        :param table: The table to clean.
        :return: The cleaned table.
        """
        array_table = np.array(table)
        non_empty_columns = ~np.all(pd.isnull(array_table), axis=0)
        return array_table[:, non_empty_columns].tolist()

    def parse_all_tables(self) -> None:
        """
        Extract tables from the Excel sheet using cell borders as indicators.
        """
        sheet_rows = list(self.sheet.iter_rows())
        border_flags = np.zeros((len(sheet_rows), len(sheet_rows[0])))

        def find_table(start_row: int, start_col: int) -> List[List[Any]]:
            header = sheet_rows[start_row]
            table_width = len(header) - 1
            for col in range(start_col, len(header)):
                if not self.cell_has_border(header[col]):
                    table_width = col - 1
                    break

            extracted_table = []
            for row in range(start_row, len(sheet_rows)):
                if not self.cell_has_border(sheet_rows[row][start_col]):
                    break
                row_values = []
                for col in range(start_col, table_width + 1):
                    row_values.append(sheet_rows[row][col].value)
                    border_flags[row][col] = 1
                if any(row_values):
                    extracted_table.append(row_values)

            if extracted_table:
                extracted_table = self.remove_empty_cols(extracted_table)
            return extracted_table

        for row_index, row in enumerate(sheet_rows):
            for col_index, cell in enumerate(row):
                if border_flags[row_index][col_index] == 0 and self.cell_has_border(cell):
                    table = find_table(row_index, col_index)
                    if table:
                        self.extracted_tables.append(table)

    def to_dataframes(self) -> List[pd.DataFrame]:
        """
        Convert the extracted tables to pandas DataFrames.

        :return: List of DataFrames.
        """
        return [pd.DataFrame(table) for table in self.extracted_tables]

    def display_tables(self) -> None:
        """
        Print the tables as pandas DataFrames.
        """
        for idx, dataframe in enumerate(self.to_dataframes()):
            print(f"Table {idx + 1}:")
            print(dataframe)
            print("\n")
            
    def serialize_tables(self) -> str:
        """
        Serialize the extracted tables to a JSON string.

        :return: JSON string containing all tables.
        """
        dataframes = self.to_dataframes()
        tables_json = [df.to_json(orient='split') for df in dataframes]
        return json.dumps(tables_json)
